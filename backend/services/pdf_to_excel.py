# backend/services/pdf_to_excel.py

import pdfplumber  # Ensure pdfplumber is imported
from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # Ensure get_column_letter is imported
from pathlib import Path
from backend.utils.file_utils import encode_file_to_base64
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def convert_pdf_to_excel(pdf_path: str, out_path: str, original_name: str):
    """
    Convert a PDF file to Excel using pdfplumber for both text and table extraction.
    Returns a dictionary containing success, filename, and base64-encoded file.
    """
    try:
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"Content_{Path(pdf_path).stem}"

        # Extract text and table data using pdfplumber
        text_content = extract_text_from_pdf(pdf_path)
        table_data = extract_tables_from_pdf(pdf_path)

        # Write extracted text content to Excel
        current_row = write_text_to_excel(ws, text_content)

        # Write extracted table data to Excel (SAFE)
        if table_data:
            write_tables_to_excel(ws, table_data, current_row)
        else:
            logging.info("No tables found - skipping table writing step")

        # Save the Excel file to the output path (ALWAYS SAFE)
        wb.save(out_path)
        logging.info(f"Excel file saved to {out_path}")

        # Encode the output Excel file to base64 for frontend use
        encoded_file = encode_file_to_base64(out_path)

        return {
            "success": True,
            "filename": f"{Path(original_name).stem}.xlsx",
            "file": encoded_file
        }

    except Exception as e:
        logging.exception("Error in convert_pdf_to_excel")
        return {
            "success": False,
            "message": f"An error occurred: {e}"
        }


def extract_text_from_pdf(pdf_path: str):
    """
    Extract text from the PDF using pdfplumber.
    The text will be placed into rows in the Excel file.
    """
    text_content = ""

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text_content += page.extract_text() or ""

    logging.info(f"Extracted {len(text_content)} characters of text from {pdf_path}")
    return text_content


def extract_tables_from_pdf(pdf_path: str):
    """
    Extract tables from all pages of a PDF file using pdfplumber.
    """
    table_data = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()

            if tables:
                logging.info(f"Found {len(tables)} table(s) on page {page.page_number}")
                table_data.extend(tables)
            else:
                logging.info(f"No tables found on page {page.page_number}")

    return table_data


def write_text_to_excel(ws, text_content):
    """
    Write extracted text content to Excel worksheet.
    Each line becomes a row.
    """
    current_row = 1

    for line in text_content.split("\n"):
        if line.strip():
            ws.cell(row=current_row, column=1, value=line)
            current_row += 1

    ws.column_dimensions['A'].width = 50
    logging.info(f"Written text content to Excel, {current_row - 1} rows")

    return current_row + 2  # leave spacing before tables


def write_tables_to_excel(ws, table_data, starting_row):
    """
    Write extracted table data safely into Excel.
    """
    row_offset = starting_row

    for table_index, table in enumerate(table_data):
        if not table:
            continue

        for r, row in enumerate(table):
            if not row:
                continue

            for c, cell in enumerate(row):
                ws.cell(
                    row=row_offset + r,
                    column=c + 1,
                    value=cell if cell is not None else ""
                )

        row_offset += len(table) + 2

    # ---------------- SAFE COLUMN WIDTH LOGIC ----------------
    max_cols = 0

    for table in table_data:
        for row in table:
            if row:
                max_cols = max(max_cols, len(row))

    if max_cols == 0:
        logging.info("No table columns found - skipping column width adjustment")
        return

    for col in range(1, max_cols + 1):
        max_length = 0

        for row in ws.iter_rows(
            min_row=starting_row,
            max_row=row_offset,
            min_col=col,
            max_col=col
        ):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    logging.info(f"Written {len(table_data)} tables to Excel")